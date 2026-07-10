# -*- coding: utf-8 -*-
"""
Argus — Raio-X dos CONCORRENTES (administradores fiduciários / DTVMs).

Para entender quem administra fundos hoje, quanto e de que tipo. Gera concorrentes.xlsx
com, por administrador: nº de fundos administrados, PL sob administração, quebra por tipo
(FIF/FIDC/FIP/FII), tipo de instituição (DTVM/CTVM/Banco/Gestora), se é IF, controle
acionário, Ativo Total e contato. Ordenado por nº de fundos (maiores concorrentes no topo).

Fonte: CVM registro_fundo_classe.zip -> registro_fundo.csv (Res.175, fundos ATIVOS)
       + cad_adm_cart (dados da empresa administradora)
       + BCB Instituicoes_em_funcionamento (segmento/tipo) e IFDATA (Ativo Total).

Rodar: python concorrentes.py   Saída: concorrentes.xlsx
"""
import os, io, csv, json, zipfile, collections, urllib.request

BASE = os.path.dirname(os.path.abspath(__file__)); DADOS = os.path.join(BASE, "dados")
os.makedirs(DADOS, exist_ok=True)
RFC = "https://dados.cvm.gov.br/dados/FI/CAD/DADOS/registro_fundo_classe.zip"
CVM_ZIP = "https://dados.cvm.gov.br/dados/ADM_CART/CAD/DADOS/cad_adm_cart.zip"
UA = {"User-Agent": "Mozilla/5.0 (Argus)"}

def http_get(url, timeout=120):
    with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=timeout) as r:
        return r.read()

def so_dig(s): return "".join(c for c in (s or "") if c.isdigit())

def num(s):
    s = (s or "").strip()
    if not s: return 0.0
    try: return float(s)
    except Exception: pass
    try: return float(s.replace(".", "").replace(",", "."))
    except Exception: return 0.0

def baixar(url, nome):
    dest = os.path.join(DADOS, nome)
    if not os.path.exists(dest) or os.path.getsize(dest) < 5000:
        print(f"  baixando {nome}..."); open(dest, "wb").write(http_get(url))
    return dest

def bucket(tp):
    tp = (tp or "").upper()
    if "FIDC" in tp: return "FIDC"
    if "FIP" in tp: return "FIP"
    if "FII" in tp or "IMOBILI" in tp: return "FII"
    if "FIF" in tp or tp == "FI" or "FINANCEIRO" in tp: return "FIF"
    return "Outros"

def tipo_por_nome(n):
    n = n.upper()
    if "DISTRIBUIDORA" in n or " DTVM" in n or "DTVM " in n: return "DTVM"
    if "CORRETORA" in n or "CTVM" in n or "CCTVM" in n or "CCVM" in n: return "CTVM"
    if n.startswith("BANCO") or "CAIXA ECON" in n: return "Banco"
    if "TRUST" in n or "FIDUCI" in n: return "Trust/serv. fiduciário"
    if "GEST" in n or "ASSET" in n or "CAPITAL" in n or "INVEST" in n: return "Gestora/asset"
    return "Outro"

def main():
    print("Argus — raio-X dos concorrentes...")
    # 1) fundos ativos por administrador
    z = zipfile.ZipFile(baixar(RFC, "registro_fundo_classe.zip"))
    fundos = list(csv.DictReader(io.StringIO(z.read("registro_fundo.csv").decode("latin-1")), delimiter=";"))
    ativos = [r for r in fundos if "FUNCIONAMENTO NORMAL" in (r.get("Situacao") or "").upper()]
    print(f"  fundos ativos (Res.175): {len(ativos)}")
    agg = collections.defaultdict(lambda: {"n": 0, "pl": 0.0, "tipos": collections.Counter(),
                                           "nomes": collections.Counter()})
    for r in ativos:
        raiz = so_dig(r.get("CNPJ_Administrador"))[:8]
        if not raiz: continue
        a = agg[raiz]; a["n"] += 1; a["pl"] += num(r.get("Patrimonio_Liquido"))
        a["tipos"][bucket(r.get("Tipo_Fundo"))] += 1
        if (r.get("Administrador") or "").strip(): a["nomes"][r["Administrador"].strip()] += 1

    # 2) dados da empresa (CVM adm_cart)
    zc = zipfile.ZipFile(baixar(CVM_ZIP, "cad_adm_cart.zip"))
    cvm = {}
    for r in csv.DictReader(io.StringIO(zc.read("cad_adm_cart_pj.csv").decode("latin-1")), delimiter=";"):
        raiz = so_dig(r["CNPJ"])[:8]
        cvm[raiz] = {"denom": r["DENOM_SOCIAL"].strip(), "categ": r["CATEG_REG"].strip(),
                     "if": r["SUBCATEG_REG"].strip() == "Instituição Financeira",
                     "controle": r["CONTROLE_ACIONARIO"].strip(), "uf": r["UF"].strip(),
                     "mun": r["MUN"].strip().title(), "dt": r["DT_REG"].strip(),
                     "email": r["EMAIL"].strip(), "tel": (r["DDD"].strip() + " " + r["TEL"].strip()).strip(),
                     "site": r["SITE_ADMIN"].strip()}

    # 3) BCB segmento + ativo total
    bcb = {}
    p = os.path.join(DADOS, "bcb_instituicoes.json")
    if os.path.exists(p):
        for x in json.load(open(p, encoding="utf-8")):
            bcb[so_dig(x["cnpj"]).zfill(8)] = x
    ativo = {}
    pa = os.path.join(DADOS, "ativos.json")
    if os.path.exists(pa):
        ativo = json.load(open(pa, encoding="utf-8"))

    linhas = []
    for raiz, a in agg.items():
        c = cvm.get(raiz, {})
        nome = c.get("denom") or (a["nomes"].most_common(1)[0][0] if a["nomes"] else raiz)
        seg = bcb.get(raiz, {}).get("seg", "")
        tipo = seg if seg else tipo_por_nome(nome)
        linhas.append({
            "Administrador": nome, "Tipo/Segmento": tipo,
            "É instituição financeira?": "sim" if (c.get("if") or raiz in bcb) else "",
            "Controle": c.get("controle", ""), "UF": c.get("uf") or bcb.get(raiz, {}).get("uf", ""),
            "Município": c.get("mun", ""), "Nº fundos": a["n"],
            "PL sob adm. (R$)": round(a["pl"], 0),
            "Nº FIF": a["tipos"].get("FIF", 0), "Nº FIDC": a["tipos"].get("FIDC", 0),
            "Nº FIP": a["tipos"].get("FIP", 0), "Nº FII/outros": a["tipos"].get("FII", 0) + a["tipos"].get("Outros", 0),
            "Foco": a["tipos"].most_common(1)[0][0] if a["tipos"] else "",
            "Ativo Total (R$)": ativo.get(raiz, "") or "", "Registro adm.": c.get("dt", ""),
            "Telefone": c.get("tel", ""), "E-mail": c.get("email", ""), "Site": c.get("site", ""),
            "CNPJ raiz": raiz,
        })
    linhas.sort(key=lambda r: -r["Nº fundos"])

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    COLS = ["Administrador", "Tipo/Segmento", "É instituição financeira?", "Controle", "UF", "Município",
            "Nº fundos", "PL sob adm. (R$)", "Nº FIF", "Nº FIDC", "Nº FIP", "Nº FII/outros", "Foco",
            "Ativo Total (R$)", "Registro adm.", "Telefone", "E-mail", "Site", "CNPJ raiz"]
    W = {"Administrador": 46, "Tipo/Segmento": 28, "É instituição financeira?": 12, "Controle": 14,
         "UF": 5, "Município": 16, "Nº fundos": 10, "PL sob adm. (R$)": 20, "Foco": 10,
         "Ativo Total (R$)": 18, "E-mail": 30, "Site": 26}
    wb = Workbook(); HF = PatternFill("solid", fgColor="6A50AC"); HFONT = Font(bold=True, color="FFFFFF", size=10)
    def aba(ws, titulo, dados):
        ws.title = titulo; ws.append(COLS)
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(1, c); cell.fill = HF; cell.font = HFONT; cell.alignment = Alignment(vertical="center")
        for r in dados:
            ws.append([r.get(k, "") for k in COLS])
            for cn in ("PL sob adm. (R$)", "Ativo Total (R$)"):
                ws.cell(ws.max_row, COLS.index(cn) + 1).number_format = "#,##0"
        for i, col in enumerate(COLS, 1): ws.column_dimensions[get_column_letter(i)].width = W.get(col, 11)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    aba(wb.active, "Concorrentes", linhas)
    aba(wb.create_sheet(), "Só DTVMs", [r for r in linhas if "DTVM" in r["Tipo/Segmento"] or "Distribuidora" in r["Tipo/Segmento"]])
    ms = wb.create_sheet("Metodologia")
    for row in [
        ["Argus — Concorrentes (administradores fiduciários). Gerado por concorrentes.py"], [""],
        ["Universo: TODO administrador com fundos ATIVOS (CVM registro_fundo, Res.175)."],
        ["Nº fundos = fundos ativos administrados; PL sob adm. = soma do PL desses fundos."],
        ["Quebra FIF/FIDC/FIP/FII mostra o FOCO (quem é de crédito estruturado vs financeiro)."],
        ["Aba 'Só DTVMs' = filtro pelas distribuidoras. Autofiltro ligado — filtre à vontade."], [""],
        ["Leitura p/ a tese: os grandes concentram; DTVMs pequenas focam FIDC. Confirme onde estão"],
        ["os fundos pequenos genéricos (FIF pequenos) — é o espaço que a Argus quer atacar."], [""],
        ["Fontes: CVM registro_fundo_classe.zip + cad_adm_cart.zip; BCB Instituicoes + IFDATA (Ativo)."],
    ]:
        ms.append(row)
    ms.column_dimensions["A"].width = 96
    out = os.path.join(BASE, "concorrentes.xlsx")
    try: wb.save(out)
    except PermissionError:
        out = os.path.join(BASE, "concorrentes_NOVO.xlsx"); wb.save(out); print("  (arquivo aberto — salvei _NOVO)")

    print(f"\n  administradores com fundos ativos: {len(linhas)}")
    ndtvm = sum(1 for r in linhas if "DTVM" in r["Tipo/Segmento"] or "Distribuidora" in r["Tipo/Segmento"])
    print(f"  dos quais DTVMs: {ndtvm}")
    print(f"  Excel: {out}")
    print("\n  Top 12 por nº de fundos:")
    for r in linhas[:12]:
        pl = r["PL sob adm. (R$)"] / 1e9
        print(f"    {r['Administrador'][:40]:40} | {r['Tipo/Segmento'][:22]:22} | {r['Nº fundos']:5} fundos | PL R$ {pl:,.1f} bi | foco {r['Foco']}")

if __name__ == "__main__":
    main()
