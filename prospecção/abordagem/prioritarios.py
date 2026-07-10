# -*- coding: utf-8 -*-
"""
Argus — gera a short-list de alvos prioritários (CRM starter) a partir de bancos_alvo.xlsx.
Filtra Alvo A (banco/CTVM elegível direto) com Perfil = 'Independente nacional', ordena por
Ativo Total (menor primeiro), e acrescenta colunas de acompanhamento (decisor, status...).

Rodar: python prioritarios.py   Saída: alvos_prioritarios.xlsx (nesta pasta)
"""
import os, csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "..", "bancos_alvo.csv")   # lê do CSV (não trava como o .xlsx aberto)

def _num(s):
    try: return float(s)
    except Exception: return None

with open(SRC, encoding="utf-8-sig") as f:
    src = [r for r in csv.DictReader(f)
           if r.get("Elegibilidade", "").startswith("Elegível") and r.get("Perfil") == "Independente nacional"]

rows = []
for r in src:
    ativo = _num(r.get("Ativo Total (R$)"))
    gestor = r.get("Gestor CVM?") == "sim"
    custod = r.get("Custodiante?") == "sim"
    sinal = " + ".join([s for s, ok in (("já gere fundos", gestor), ("já custodia", custod)) if ok]) or "—"
    rows.append({
        "Prioridade": "", "Instituição": r.get("Instituição"), "Segmento BCB": r.get("Segmento BCB"),
        "UF": r.get("UF"), "Município": r.get("Município"),
        "Ativo Total (R$)": ativo if ativo else "", "Início atividade": r.get("Início atividade"),
        "Já no mundo de fundos?": sinal, "Nº fundos que gere": r.get("Nº fundos que gere"),
        "Nº classes que custodia": r.get("Nº classes que custodia"),
        "One-pager": "Banco/CTVM", "E-mail": r.get("E-mail"), "Telefone": r.get("Telefone"),
        "Site": r.get("Site"),
        # colunas de acompanhamento (preencher)
        "Decisor": "", "Cargo": "", "LinkedIn": "", "Status": "", "Última ação": "",
        "Próximo passo": "", "Notas": "",
        "_at": ativo if isinstance(ativo, (int, float)) else 10**15,
        "_warm": not (gestor or custod),
    })

# ordena: menor Ativo Total primeiro (sem ativo vai pro fim)
rows.sort(key=lambda r: r["_at"])

OUT_COLS = ["Prioridade", "Instituição", "Segmento BCB", "UF", "Município", "Ativo Total (R$)",
            "Início atividade", "Já no mundo de fundos?", "Nº fundos que gere", "Nº classes que custodia",
            "One-pager", "E-mail", "Telefone", "Site",
            "Decisor", "Cargo", "LinkedIn", "Status", "Última ação", "Próximo passo", "Notas"]
W = {"Instituição": 42, "Segmento BCB": 26, "Município": 16, "Ativo Total (R$)": 16,
     "Já no mundo de fundos?": 22, "One-pager": 12, "E-mail": 30, "Telefone": 15, "Site": 26,
     "Decisor": 22, "Cargo": 18, "LinkedIn": 24, "Status": 14, "Última ação": 16, "Próximo passo": 20, "Notas": 26}

out = openpyxl.Workbook(); wsx = out.active; wsx.title = "Prioritários"
HF = PatternFill("solid", fgColor="6A50AC"); HFONT = Font(bold=True, color="FFFFFF", size=10)
ACC = PatternFill("solid", fgColor="EFEBFA")  # destaque suave (roxo claro) nas colunas de preencher
wsx.append(OUT_COLS)
for c in range(1, len(OUT_COLS) + 1):
    cell = wsx.cell(1, c); cell.fill = HF; cell.font = HFONT; cell.alignment = Alignment(vertical="center")
fill_from = OUT_COLS.index("Decisor")
for r in rows:
    wsx.append([r.get(k, "") for k in OUT_COLS])
    wsx.cell(wsx.max_row, OUT_COLS.index("Ativo Total (R$)") + 1).number_format = "#,##0"
    for ci in range(fill_from, len(OUT_COLS)):
        wsx.cell(wsx.max_row, ci + 1).fill = ACC
for i, col in enumerate(OUT_COLS, 1):
    wsx.column_dimensions[get_column_letter(i)].width = W.get(col, 12)
wsx.freeze_panes = "B2"; wsx.auto_filter.ref = f"A1:{get_column_letter(len(OUT_COLS))}{wsx.max_row}"

path = os.path.join(BASE, "alvos_prioritarios.xlsx")
try: out.save(path)
except PermissionError:
    path = os.path.join(BASE, "alvos_prioritarios_NOVO.xlsx"); out.save(path)
print(f"gerado: {path} | {len(rows)} prioritários (independente nacional, elegível direto)")
print("top 12 (menor Ativo Total):")
for r in rows[:12]:
    at = f"{r['_at']/1e6:,.0f} mi" if r['_at'] < 10**15 else "?"
    print(f"   {r['Instituição'][:40]:40} | {r['UF']:2} | R$ {at:>8} | {r['Já no mundo de fundos?']}")
