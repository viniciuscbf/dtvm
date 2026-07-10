# -*- coding: utf-8 -*-
"""
Argus — AUDITORES INDEPENDENTES qualificados para auditar fundos (candidatas baratas).

Gera auditores_alvo.xlsx com as firmas de auditoria registradas na CVM, quantos fundos cada
uma audita hoje (proxy de experiência), tier (Big Four / rede internacional / boutique), CNPJ,
cidade/UF, situação e contato. A ideia é achar as BOUTIQUES que já auditam fundos — qualificadas,
sem prêmio Big Four e famintas por volume — para a Argus negociar auditoria em lote.

Fontes (dados abertos CVM, atualizados até o último dia útil):
  - Cadastro de auditores:  dados/AUDITOR/CAD/DADOS/cad_auditor.zip -> cad_auditor_pj.csv
  - Vínculo auditor->fundo: dados/FI/CAD/DADOS/cad_fi.csv (colunas AUDITOR e CNPJ_AUDITOR)
  - Contato:                BrasilAPI /api/cnpj/v1/{cnpj} (telefone/e-mail), com cache

Rodar: python auditores.py   Saída: auditores_alvo.xlsx (+ auditores_candidatos_argus.csv)
"""
import os, io, csv, json, time, zipfile, collections, urllib.request

BASE = os.path.dirname(os.path.abspath(__file__)); DADOS = os.path.join(BASE, "dados")
os.makedirs(DADOS, exist_ok=True)
AUD_ZIP = "https://dados.cvm.gov.br/dados/AUDITOR/CAD/DADOS/cad_auditor.zip"
CAD_FI = "https://dados.cvm.gov.br/dados/FI/CAD/DADOS/cad_fi.csv"
BRASILAPI = "https://brasilapi.com.br/api/cnpj/v1/{}"
UA = {"User-Agent": "Mozilla/5.0 (Argus)"}
CACHE = os.path.join(DADOS, "cache_cnpj.json")

# Big Four + redes internacionais (para separar das boutiques nacionais)
BIG4 = ["KPMG", "PRICEWATER", "ERNST", "DELOITTE"]
REDES = ["BDO", "GRANT THORNTON", "RSM", "UHY", "BAKER TILLY", "MGI", "MOORE", "MAZARS", "FORVIS",
         "CROWE", "PKF", "NEXIA", "KRESTON", "HLB", "BKR", "GRACE", "PARKER RANDALL", "RUSSELL BEDFORD"]


def http_get(url, timeout=180):
    with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=timeout) as r:
        return r.read()


def baixar(url, nome, minsize=5000):
    dest = os.path.join(DADOS, nome)
    if not os.path.exists(dest) or os.path.getsize(dest) < minsize:
        print(f"  baixando {nome}..."); open(dest, "wb").write(http_get(url))
    return dest


def so_dig(s): return "".join(c for c in (s or "") if c.isdigit())


def norm(s):
    """normaliza razão social para casar cad_fi.AUDITOR com cad_auditor_pj.DENOM_SOCIAL."""
    s = (s or "").upper()
    for w in ["LTDA", "S/S", "SOCIEDADE SIMPLES", "EPP", " ME ", "AUDITORES", "AUDITORIA",
              "INDEPENDENTES", "ASSOCIADOS", "ASSURANCE", "CIA", " E "]:
        s = s.replace(w, " ")
    return "".join(c for c in s if c.isalnum())


def tier(nome):
    u = nome.upper()
    if any(b in u for b in BIG4): return "Big Four"
    if any(b in u for b in REDES): return "Rede intl"
    return "Boutique"


def contato(cnpj14, cache):
    """telefone/e-mail via BrasilAPI, com cache em disco."""
    if not cnpj14 or len(cnpj14) != 14: return "", ""
    if cnpj14 in cache: return cache[cnpj14].get("tel", ""), cache[cnpj14].get("email", "")
    tel = em = ""
    try:
        j = json.loads(http_get(BRASILAPI.format(cnpj14), timeout=15).decode("utf-8"))
        tel = (j.get("ddd_telefone_1") or "").strip()
        em = (j.get("email") or "").strip()
        time.sleep(0.4)
    except Exception:
        pass
    cache[cnpj14] = {"tel": tel, "email": em}
    return tel, em


def main():
    print("Argus — auditores candidatos (fundos)...")
    # 1) firmas registradas (PJ) -> CNPJ/cidade/UF/situação
    z = zipfile.ZipFile(baixar(AUD_ZIP, "cad_auditor.zip"))
    pj = {}
    for r in csv.DictReader(io.StringIO(z.read("cad_auditor_pj.csv").decode("latin-1")), delimiter=";"):
        pj[norm(r["DENOM_SOCIAL"])] = {"cnpj": r["CNPJ"].strip(), "denom": r["DENOM_SOCIAL"].strip(),
                                       "mun": (r.get("MUN") or "").strip().title(), "uf": (r.get("UF") or "").strip(),
                                       "sit": (r.get("SIT") or "").strip()}
    ativas = sum(1 for v in pj.values() if "ATIVO" in v["sit"].upper())
    print(f"  firmas PJ registradas: {len(pj)} (ativas: {ativas})")

    # 2) contar fundos por auditor (coluna AUDITOR do cad_fi)
    fh = io.StringIO(open(baixar(CAD_FI, "cad_fi.csv", 1_000_000), "rb").read().decode("latin-1"))
    cnt = collections.Counter()
    for r in csv.DictReader(fh, delimiter=";"):
        a = (r.get("AUDITOR") or "").strip()
        if a: cnt[a] += 1
    print(f"  fundos com auditor informado: {sum(cnt.values())} | firmas que auditam fundos: {len(cnt)}")

    # 3) montar linhas (só quem audita fundos), casar com o cadastro PJ, classificar tier
    cache = json.load(open(CACHE, encoding="utf-8")) if os.path.exists(CACHE) else {}
    linhas = []
    for nome, n in cnt.most_common():
        t = tier(nome)
        m = pj.get(norm(nome))
        if not m:  # match aproximado por prefixo
            k = norm(nome)
            for kk, vv in pj.items():
                if kk and (kk.startswith(k[:14]) or k.startswith(kk[:14])): m = vv; break
        cnpj = so_dig(m["cnpj"]) if m else ""
        tel, em = ("", "")
        if t != "Big Four" and cnpj:  # enriquece só candidatas (economiza chamadas)
            tel, em = contato(cnpj, cache)
        linhas.append({"Firma": (m["denom"] if m else nome), "Tier": t, "Nº fundos auditados": n,
                       "CNPJ": (m["cnpj"] if m else ""), "Município": (m["mun"] if m else ""),
                       "UF": (m["uf"] if m else ""), "Situação": (m["sit"] if m else "?"),
                       "Telefone": tel, "E-mail": em})
    json.dump(cache, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False, indent=0)

    candidatas = [r for r in linhas if r["Tier"] != "Big Four"]

    # 4) Excel (cabeçalho roxo da marca)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    COLS = ["Firma", "Tier", "Nº fundos auditados", "CNPJ", "Município", "UF", "Situação", "Telefone", "E-mail"]
    W = {"Firma": 52, "Tier": 12, "Nº fundos auditados": 16, "CNPJ": 20, "Município": 18, "UF": 5,
         "Situação": 12, "Telefone": 16, "E-mail": 30}
    wb = Workbook(); HF = PatternFill("solid", fgColor="6A50AC"); HFONT = Font(bold=True, color="FFFFFF", size=10)

    def aba(ws, titulo, dados):
        ws.title = titulo; ws.append(COLS)
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(1, c); cell.fill = HF; cell.font = HFONT; cell.alignment = Alignment(vertical="center")
        for r in dados: ws.append([r.get(k, "") for k in COLS])
        for i, col in enumerate(COLS, 1): ws.column_dimensions[get_column_letter(i)].width = W.get(col, 12)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"

    aba(wb.active, "Candidatas (não-BigFour)", candidatas)
    aba(wb.create_sheet(), "Todos que auditam fundos", linhas)
    ms = wb.create_sheet("Metodologia")
    big_funds = sum(r["Nº fundos auditados"] for r in linhas if r["Tier"] == "Big Four")
    tot_funds = sum(r["Nº fundos auditados"] for r in linhas)
    for row in [
        ["Argus — Auditores candidatos para fundos. Gerado por auditores.py"], [""],
        [f"Firmas que auditam fundos: {len(linhas)} | candidatas (não-Big-Four): {len(candidatas)}"],
        [f"Fundos com auditor informado: {tot_funds} | Big Four: {big_funds} ({100*big_funds/tot_funds:.1f}%)"], [""],
        ["Alvo: BOUTIQUES que já auditam fundos (qualificadas, sem prêmio Big Four, apetite por volume)."],
        ["Nº fundos auditados = proxy de experiência (campo AUDITOR autodeclarado no cad_fi, ~41% preenchido)."],
        ["Preço-alvo defensável: R$ 3-5 mil/fundo (mercado ~R$ 11 mil). Ver guia_auditor_independente.md §11."],
        ["Rodízio: firma sai do MESMO fundo após 5 exercícios (3 de intervalo). Manter 2 firmas em rodízio."], [""],
        ["Fontes: CVM cad_auditor.zip + cad_fi.csv (campo AUDITOR); contato via BrasilAPI (Receita)."],
        ["Ressalva: registrada != topa preço baixo; telefones podem estar desatualizados — validar."],
    ]:
        ms.append(row)
    ms.column_dimensions["A"].width = 100

    out = os.path.join(BASE, "auditores_alvo.xlsx")
    try: wb.save(out)
    except PermissionError:
        out = os.path.join(BASE, "auditores_alvo_NOVO.xlsx"); wb.save(out); print("  (arquivo aberto — salvei _NOVO)")

    # CSV espelho (Excel-friendly, na raiz do projeto)
    csvout = os.path.join(os.path.dirname(BASE), "auditores_candidatos_argus.csv")
    with open(csvout, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";"); w.writerow(COLS)
        for r in candidatas: w.writerow([r.get(k, "") for k in COLS])

    print(f"\n  candidatas (não-Big-Four que auditam fundos): {len(candidatas)}")
    print(f"  Big Four: {100*big_funds/tot_funds:.1f}% dos fundos auditados")
    print(f"  Excel: {out}\n  CSV:   {csvout}")
    print("\n  Top 12 candidatas (boutiques/redes):")
    for r in candidatas[:12]:
        print(f"    {r['Firma'][:44]:44} | {r['Tier']:9} | {r['Nº fundos auditados']:4} fundos | {r['Município']}/{r['UF']} | {r['Telefone']}")


if __name__ == "__main__":
    main()
