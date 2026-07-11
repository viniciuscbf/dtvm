# -*- coding: utf-8 -*-
"""
Argus — Prospecção de parceiros (versão completa, com colunas para filtrar no Excel).

Alvo: instituição que possa hospedar a administradora fiduciária de fundos pequenos.
Regras: NÃO pode ser administrador fiduciário (inegociável); DTVM excluída; CTVM pode;
custodiante pode (flag). Amplia p/ não-bancárias com capital (SCD/IP/financeira/câmbio...).

Em vez de FILTRAR fora (estrangeiro/captivo/estatal/fintech...), tudo isso vira COLUNA
(`Perfil`) — você filtra/desfiltra no próprio Excel. Porte = **Ativo Total** (BCB IF.data),
não capital social.

Fontes:
  BCB  Instituicoes_em_funcionamento (SedesBancoComMultCE + SedesSociedades) — nome, segmento, contato
  BCB  IFDATA IfDataValores (TipoInstituicao=2, Relatório Resumo) — Ativo Total por CNPJ (raiz)
  CVM  cad_adm_cart.zip (adm fiduciário / gestor) + tabecus.asp (custodiantes)

Rodar: python build_lista.py   Saída: bancos_alvo.xlsx + .csv
"""
import os, io, csv, json, zipfile, re, collections, unicodedata, urllib.request

BASE = os.path.dirname(os.path.abspath(__file__)); DADOS = os.path.join(BASE, "dados")
os.makedirs(DADOS, exist_ok=True)

BCB_SOC = "https://olinda.bcb.gov.br/olinda/servico/Instituicoes_em_funcionamento/versao/v1/odata/SedesSociedades?$format=json"
BCB_BCO = "https://olinda.bcb.gov.br/olinda/servico/Instituicoes_em_funcionamento/versao/v1/odata/SedesBancoComMultCE?$format=json"
IFDATA_VAL = ("https://olinda.bcb.gov.br/olinda/servico/IFDATA/versao/v1/odata/"
              "IfDataValores(AnoMes=@AnoMes,TipoInstituicao=@TipoInstituicao,Relatorio=@Relatorio)"
              "?@AnoMes={ym}&@TipoInstituicao=2&@Relatorio='1'&$format=json&$top=40000")
CVM_ZIP = "https://dados.cvm.gov.br/dados/ADM_CART/CAD/DADOS/cad_adm_cart.zip"
CVM_CUS = "https://sistemas.cvm.gov.br/asp/cvmwww/invnres/tabecus.asp"
IFDATA_CAD = "https://olinda.bcb.gov.br/olinda/servico/IFDATA/versao/v1/odata/IfDataCadastro(AnoMes=@AnoMes)?@AnoMes={ym}&$format=json&$top=8000"
RFC_URL = "https://dados.cvm.gov.br/dados/FI/CAD/DADOS/registro_fundo_classe.zip"
UA = {"User-Agent": "Mozilla/5.0 (Argus prospeccao)"}

# ---- listas para a COLUNA Perfil (não filtram nada; só rotulam) ----
ESTATAL = ["BANCO DO BRASIL", "BNDES", "BANCO NACIONAL DE DESENV", "CAIXA ECON", "BANCO DA AMAZONIA",
    "BANCO DO NORDESTE", "BANCO DO ESTADO", "BANESTES", "BANESE", "BANPARA", "BANPARÁ", "BANRISUL",
    "BRB", "BANCO DE BRASILIA", "BANCO REGIONAL DE"]
ESTRANGEIRO = ["CITI", "JPMORGAN", "J.P. MORGAN", "CREDIT SUISSE", "UBS", "MORGAN STANLEY", "GOLDMAN",
    "BNP PARIBAS", "DEUTSCHE", "BANK OF AMERICA", "SCOTIABANK", "BNY MELLON", "STATE STREET",
    "NORTHERN TRUST", "MUFG", "MIZUHO", "SUMITOMO", "WOORI", "KEB HANA", "KDB", "BANK OF CHINA",
    "ICBC", "ANDBANK", "BOCOM", "DE LA NACION", "DE LA PROVINCIA", "HSBC", "STANDARD", "RABOBANK",
    "DE LAGE LANDEN", "WESTERN UNION", "TRAVELEX", "MONEYCORP", "INTESA", "CHINA CONSTRUCTION",
    "SUMITOMO MITSUI", "CARGILL", "BANK OF", "COMMERZBANK", "NATIXIS", "SOCIETE GENERALE"]
CAPTIVO = ["STELLANTIS", "PACCAR", "JOHN DEERE", "XCMG", "KOMATSU", "TOYOTA", "HONDA", "VOLKSWAGEN",
    "VOLKSWAGE", "MERCEDES", "VOLVO", "SCANIA", "CNH INDUSTRIAL", "CATERPILLAR", "FORD",
    "GENERAL MOTORS", "HYUNDAI", "YAMAHA", "MONEO", "RANDON", "IVECO", "PSA", "RCI ", "MAN ",
    "AGCO", "DAIMLER", "BANCO GM"]
GRANDE = ["BRADESCO", "ITAU", "ITAÚ", "SANTANDER", "BTG", "SAFRA", "NUBANK", "NU PAGAMENTOS",
    "BANCO INTER", "C6 ", "BANCO PAN", "BANCO ORIGINAL", "BANCO BMG", "BANCO VOTORANTIM", "BANCO BV",
    "XP ", "BANCO GENIAL", "BANCO DAYCOVAL", "BANCO MODAL", "MERCADO PAGO", "MERCADO CREDITO",
    "PAGSEGURO", "PICPAY", "STONE", "AGIBANK", "BANCO DIGIO", "CREFISA", "MIDWAY", "WILL FINANCEIRA"]

def http_get(url, timeout=90):
    with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=timeout) as r:
        return r.read()

def so_dig(s): return "".join(c for c in (s or "") if c.isdigit())

def num(s):
    s = (s or "").strip()
    if not s: return 0.0
    try: return float(s)
    except Exception: pass
    try: return float(s.replace(".", "").replace(",", "."))
    except Exception: return 0.0

def cnpj_dv(base12):
    def d(nums, pesos):
        s = sum(int(n) * p for n, p in zip(nums, pesos)); r = s % 11
        return "0" if r < 2 else str(11 - r)
    p1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]; d1 = d(base12, p1)
    return d1 + d(base12 + d1, [6] + p1)

def fmt_cnpj(raiz8):
    b = raiz8.zfill(8) + "0001"; c = b + cnpj_dv(b)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"

def _has(nome, lst):
    n = (nome or "").upper(); return any(t in n for t in lst)

# --- detecção de "mesmo grupo de uma DTVM/adm.fiduciário" pela MARCA (1º nome não-genérico) ---
def _norm(s):
    return unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode().lower()

SKIP_TOK = {"banco", "bco", "agencia", "sociedade", "caixa", "cooperativa", "coop", "companhia",
            "cia", "distribuidora", "corretora", "ctvm", "dtvm", "cctvm", "ccvm", "cvmc",
            "associacao", "de", "da", "do", "dos", "das", "e", "s", "sa"}
GENERIC_BRAND = {"nova", "novo", "br", "brasil", "uniao", "central", "geral", "global", "grupo",
                 "primeiro", "banco", "capital", "invest", "investimentos"}

def marca(nome):
    for t in re.split(r"[\s/.,\-()]+", _norm(nome)):
        if t and t not in SKIP_TOK and len(t) >= 3:
            return t
    return ""

# ------------------------------------------------------ dados
def universo_bcb():
    cache = os.path.join(DADOS, "bcb_instituicoes.json")
    if os.path.exists(cache): return json.load(open(cache, encoding="utf-8"))
    out = []
    for url, fonte in ((BCB_BCO, "banco"), (BCB_SOC, "sociedade")):
        v = json.loads(http_get(url))["value"]
        for x in v:
            out.append({"cnpj": so_dig(x.get("CNPJ")), "nome": (x.get("NOME_INSTITUICAO") or "").strip(),
                        "seg": (x.get("SEGMENTO") or "").strip(), "uf": x.get("UF") or "",
                        "mun": (x.get("MUNICIPIO") or "").title(),
                        "tel": ((x.get("DDD") or "").strip() + " " + (x.get("TELEFONE") or "").strip()).strip(),
                        "email": (x.get("E_MAIL") or "").strip(), "site": (x.get("SITIO_NA_INTERNET") or "").strip()})
        print(f"  BCB {fonte}: {len(v)}")
    json.dump(out, open(cache, "w", encoding="utf-8"), ensure_ascii=False)
    return out

# Colunas de porte do IFDATA (Relatório '1') -> chave interna. As de ESCALA entram no teto;
# Lucro Líquido é resultado (pode ser negativo) — vira coluna informativa, NÃO entra no filtro.
PORTE_COLS = {
    "Ativo Total": "ativo", "Passivo Exigível": "passivo", "Patrimônio Líquido": "pl",
    "Captações": "captacoes", "Carteira de Crédito": "carteira",
    "Títulos e Valores Mobiliários": "tvm", "Lucro Líquido": "lucro",
}
ESCALA_KEYS = ("ativo", "passivo", "pl", "captacoes", "carteira", "tvm")   # métricas de tamanho da estrutura
TETO = 400_000_000   # R$ 400 mi: nenhuma métrica de escala pode passar disso (regra do usuário)
# Rede de segurança p/ porte NÃO reportado (CNPJ-raiz não casa com IFDATA): perfis sabidamente grandes
# não podem entrar no alvo só porque o dado veio em branco (senão JP Morgan/BB/XP escapam).
BIG_PERFIS = {"Grande/varejo", "Estrangeiro", "Estatal/público", "Captivo (montadora/equip.)"}

def portes_bcb():
    """raiz do CNPJ -> {ativo, passivo, pl, captacoes, carteira, tvm, lucro} em R$ (BCB IFDATA)."""
    cache = os.path.join(DADOS, "portes.json")
    if os.path.exists(cache): return json.load(open(cache, encoding="utf-8"))
    mp = {}
    for ym in ("202503", "202412", "202409"):
        try: v = json.loads(http_get(IFDATA_VAL.format(ym=ym)))["value"]
        except Exception as e: print("  IFDATA valores falhou", ym, e); continue
        for r in v:
            k = PORTE_COLS.get(r.get("NomeColuna"))
            if k:
                mp.setdefault(str(r["CodInst"]).zfill(8), {})[k] = r["Saldo"]
        if mp: print(f"  IFDATA porte ({len(PORTE_COLS)} colunas): {len(mp)} instituições (base {ym})"); break
    json.dump(mp, open(cache, "w", encoding="utf-8"))
    return mp

def escala(p):
    """maior métrica de escala conhecida (0 se nada reportado)."""
    return max([p.get(k) or 0 for k in ESCALA_KEYS], default=0)

def cvm_flags():
    dest = os.path.join(DADOS, "cad_adm_cart.zip")
    if not os.path.exists(dest) or os.path.getsize(dest) < 10000:
        open(dest, "wb").write(http_get(CVM_ZIP))
    rows = list(csv.DictReader(io.StringIO(zipfile.ZipFile(dest).read("cad_adm_cart_pj.csv").decode("latin-1")), delimiter=";"))
    fiduc, gestor = set(), set(); fiduc_nomes = []
    for r in rows:
        if r["SIT"].strip() != "EM FUNCIONAMENTO NORMAL": continue
        raiz = so_dig(r["CNPJ"])[:8]
        if "Administrador Fiduci" in r["CATEG_REG"]:
            fiduc.add(raiz); fiduc_nomes.append(r["DENOM_SOCIAL"].strip())
        if "Gestor" in r["CATEG_REG"]: gestor.add(raiz)
    try:
        custod = {m[-14:][:8] for m in re.findall(r"\d{14,15}", http_get(CVM_CUS).decode("latin-1"))}
    except Exception: custod = set()
    return fiduc, gestor, custod, fiduc_nomes

def _rfc():
    dest = os.path.join(DADOS, "registro_fundo_classe.zip")
    if not os.path.exists(dest) or os.path.getsize(dest) < 5000:
        print("  baixando registro_fundo_classe..."); open(dest, "wb").write(http_get(RFC_URL))
    return zipfile.ZipFile(dest)

def fundos_geridos():          # nº de fundos ativos por gestor (raiz do CNPJ)
    cache = os.path.join(DADOS, "gestor_cont.json")
    if os.path.exists(cache): return json.load(open(cache))
    c = collections.Counter()
    for r in csv.DictReader(io.StringIO(_rfc().read("registro_fundo.csv").decode("latin-1")), delimiter=";"):
        if "FUNCIONAMENTO NORMAL" not in (r.get("Situacao") or "").upper(): continue
        g = so_dig(r.get("CPF_CNPJ_Gestor"))
        if len(g) == 14: c[g[:8]] += 1
    d = dict(c); json.dump(d, open(cache, "w")); return d

def classes_custodiadas():     # (nº classes, PL) por custodiante (raiz do CNPJ)
    cache = os.path.join(DADOS, "custod_cont.json")
    if os.path.exists(cache): return json.load(open(cache))
    n = collections.Counter(); pl = collections.defaultdict(float)
    for r in csv.DictReader(io.StringIO(_rfc().read("registro_classe.csv").decode("latin-1")), delimiter=";"):
        if "FUNCIONAMENTO NORMAL" not in (r.get("Situacao") or "").upper(): continue
        cu = so_dig(r.get("CNPJ_Custodiante"))
        if len(cu) == 14: n[cu[:8]] += 1; pl[cu[:8]] += num(r.get("Patrimonio_Liquido"))
    d = {k: [n[k], round(pl[k], 0)] for k in n}; json.dump(d, open(cache, "w")); return d

def inicio_atividade():        # ano de início de atividade por raiz de CNPJ (BCB IFDATA)
    cache = os.path.join(DADOS, "inicio_ativ.json")
    if os.path.exists(cache): return json.load(open(cache))
    mp = {}
    for ym in ("202503", "202412"):
        try: v = json.loads(http_get(IFDATA_CAD.format(ym=ym)))["value"]
        except Exception: continue
        for it in v:
            raiz = so_dig(str(it.get("CnpjInstituicaoLider") or "")).zfill(8)[:8]
            y = str(it.get("DataInicioAtividade") or "")[:4]
            if raiz and y.isdigit() and 1900 <= int(y) <= 2026 and raiz not in mp:
                mp[raiz] = y   # ignora sentinela "1800" do IFDATA (desconhecido)
        if mp: break
    json.dump(mp, open(cache, "w")); return mp

# ------------------------------------------------------ classificação (colunas)
def perfil(nome, seg):
    if _has(nome, ESTATAL) or "Desenvolvimento" in seg or "Fomento" in seg or "BNDES" in nome.upper() or "Caixa Econ" in seg:
        return "Estatal/público"
    if _has(nome, ESTRANGEIRO): return "Estrangeiro"
    if _has(nome, CAPTIVO): return "Captivo (montadora/equip.)"
    if _has(nome, GRANDE): return "Grande/varejo"
    if "Câmbio" in seg or "Cambio" in seg: return "Câmbio"
    if "Crédito Direto" in seg: return "Fintech (SCD)"
    if "Pagamento" in seg: return "Fintech (pagamentos)"
    if "Financiamento" in seg: return "Financeira"
    if "Hipotec" in seg: return "Hipotecária"
    return "Independente nacional"

def elegibilidade(seg, adm_fiduc):
    is_dtvm = "Distribuidora de TVM" in seg
    is_ctvm = "Corretora de TVM" in seg
    is_cambio = "Câmbio" in seg or "Cambio" in seg
    is_banco = seg.startswith("Banco") or "Caixa" in seg
    if adm_fiduc: return "Já adm. fiduciário (excluir)"
    # DTVM: excluída por DECISÃO ESTRATÉGICA (não jurídica) — já tem retaguarda/estrutura própria
    # de mercado de capitais; agregamos pouco e o risco de ela absorver a ideia e dispensar a
    # parceria é alto. (Juridicamente seria elegível às duas licenças — não importa: fora.)
    if is_dtvm: return "DTVM (excluir — estrutura própria)"
    if (is_banco and not is_cambio) or is_ctvm: return "Elegível direto (banco/CTVM)"
    # SCFI: Res. CMN 5.237/2025, art. 6º, p.u., IV — financeiras podem administrar carteiras
    # de valores mobiliários desde 1º/9/2025 (custódia continua fora — Res. CVM 32, art. 4º).
    if "Financiamento" in seg: return "Elegível adm. fiduciário (SCFI, Res. 5.237)"
    # Objeto social taxativo não comporta administração de fundos (SCD Res. 5.050 art. 7º §1º;
    # câmbio Res. 5.009 art. 2º; hipotecária Res. 2.122 só FII; IP = tese inédita) → só
    # constituindo DTVM no grupo (R$ 1,5 mi + 6-12 meses de BCB).
    if is_cambio or "Crédito Direto" in seg or "Pagamento" in seg or "Hipotec" in seg or "Fomento" in seg:
        return "Só via DTVM nova (objeto restrito)"
    return "Outro"

# ------------------------------------------------------ Excel
COLS = ["Instituição", "Segmento BCB", "Perfil", "Elegibilidade", "Grupo tem DTVM/adm.fid.?",
        "CNPJ", "UF", "Município",
        "Ativo Total (R$)", "Passivo Exigível (R$)", "Patrimônio Líquido (R$)", "Captações (R$)",
        "Carteira de Crédito (R$)", "TVM (R$)", "Maior escala (R$)", "Lucro Líquido (R$)",
        "Início atividade", "Gestor CVM?", "Nº fundos que gere",
        "Adm.fiduciário?", "Custodiante?", "Nº classes que custodia", "PL custodiado (R$)",
        "Telefone", "E-mail", "Site"]

def escrever(master, alvoA, alvoB, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook(); HF = PatternFill("solid", fgColor="6A50AC"); HFONT = Font(bold=True, color="FFFFFF", size=10)
    W = {"Instituição": 46, "Segmento BCB": 30, "Perfil": 22, "Elegibilidade": 26,
         "Grupo tem DTVM/adm.fid.?": 22, "CNPJ": 20, "UF": 5,
         "Município": 16, "Ativo Total (R$)": 18, "Passivo Exigível (R$)": 18,
         "Patrimônio Líquido (R$)": 18, "Captações (R$)": 16, "Carteira de Crédito (R$)": 18,
         "TVM (R$)": 16, "Maior escala (R$)": 16, "Lucro Líquido (R$)": 16,
         "Início atividade": 11, "Gestor CVM?": 11,
         "Nº fundos que gere": 12, "Adm.fiduciário?": 13, "Custodiante?": 12,
         "Nº classes que custodia": 13, "PL custodiado (R$)": 18, "Telefone": 15, "E-mail": 30, "Site": 26}
    money_i = [COLS.index(c) + 1 for c in ("Ativo Total (R$)", "Passivo Exigível (R$)",
               "Patrimônio Líquido (R$)", "Captações (R$)", "Carteira de Crédito (R$)", "TVM (R$)",
               "Maior escala (R$)", "Lucro Líquido (R$)", "PL custodiado (R$)")]
    def aba(ws, titulo, linhas):
        ws.title = titulo; ws.append(COLS)
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(1, c); cell.fill = HF; cell.font = HFONT; cell.alignment = Alignment(vertical="center")
        for r in linhas:
            ws.append([r.get(k, "") for k in COLS])
            for mi in money_i: ws.cell(ws.max_row, mi).number_format = "#,##0"
        for i, col in enumerate(COLS, 1): ws.column_dimensions[get_column_letter(i)].width = W[col]
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    aba(wb.active, "Master (tudo)", master)
    aba(wb.create_sheet(), "Alvos (A+B consolidado)", alvoA + alvoB)
    aba(wb.create_sheet(), "Alvo A - banco e CTVM", alvoA)
    aba(wb.create_sheet(), "Alvo B - nao-bancaria", alvoB)
    ms = wb.create_sheet("Metodologia")
    for row in [
        ["Argus — Prospecção (build_lista.py)"], [""],
        ["FILTRO INEGOCIÁVEL: não é administrador fiduciário. Excluídos do alvo: adm. fiduciário e DTVM."],
        ["CTVM entra; custodiante entra (coluna). Nada mais é filtrado — use os filtros do Excel."], [""],
        ["Coluna PERFIL rotula: Estatal/público, Estrangeiro, Captivo (montadora), Grande/varejo,"],
        ["  Câmbio, Fintech (SCD), Fintech (pagamentos), Financeira, Hipotecária, Independente nacional."],
        ["  -> filtre 'Independente nacional' no Alvo A para os alvos mais prováveis."],
        ["Coluna ELEGIBILIDADE (validada em fontes primárias — ver parceria_estrutura_juridica.md):"],
        ["  Elegível direto (banco/CTVM) = parceiro completo: adm. fiduciário (Res. 21 art. 1º §2º I) +"],
        ["    custódia própria possível (Res. 32 art. 4º) + distribuição (art. 33)."],
        ["  DTVM (excluir — estrutura própria) = decisão ESTRATÉGICA: DTVM já tem retaguarda de mercado de"],
        ["    capitais; agregamos pouco e o risco de absorver a ideia e dispensar a parceria é alto."],
        ["  Elegível adm. fiduciário (SCFI, Res. 5.237) = financeiras PODEM administrar carteiras desde"],
        ["    1º/9/2025 (art. 6º p.u. IV) — custódia sempre terceirizada (fora do rol da Res. 32)."],
        ["  Só via DTVM nova (objeto restrito) = SCD (Res. 5.050 art. 7º §1º taxativo), IP (tese inédita),"],
        ["    câmbio (Res. 5.009 objeto exclusivo), hipotecária (só FII) — exigem DTVM no grupo:"],
        ["    R$ 1,5 mi de capital + 6-12 meses de BCB. Fundo do funil."],
        ["  DTVM (excluir) | Já adm. fiduciário (excluir)."], [""],
        ["PORTE (BCB IF.data, IfDataValores TipoInstituicao=2, base ~mar/2025), em R$: Ativo Total, Passivo"],
        ["  Exigível, Patrimônio Líquido, Captações, Carteira de Crédito e TVM. Lucro Líquido é informativo."],
        ["FILTRO DE PORTE nos Alvos: nenhuma métrica de ESCALA (todas, menos o Lucro) pode passar de R$ 400 mi;"],
        ["  qualquer valor acima disso EXCLUI a instituição do alvo. A coluna 'Maior escala' mostra a que manda."],
        ["  Escala em branco (não reporta) = mantida, EXCETO se o Perfil já é grande (Estrangeiro/Grande/"],
        ["  Estatal/Captivo) — aí o branco é lacuna de dado e a instituição sai (pega JP Morgan/BB/XP sem porte)."],
        ["  O Master (aba 1) mantém TODOS, sem teto, para referência."],
        ["Alvo A ordenado por Ativo Total (menor primeiro); em branco vai para o fim."], [""],
        ["Fontes: BCB Instituicoes_em_funcionamento + IFDATA; CVM cad_adm_cart + tabecus.asp."],
        ["Sua pergunta: instituições AUTORIZADAS estão TODAS aqui. Empresas com capital que ainda não são"],
        ["instituição financeira (se autorizariam do zero) não estão em registro — só rede/internet."],
    ]:
        ms.append(row)
    ms.column_dimensions["A"].width = 100
    wb.save(path)

# ------------------------------------------------------ main
def main():
    print("Argus — prospecção (Ativo Total + colunas de perfil)...")
    fiduc, gestor, custod, fiduc_nomes = cvm_flags(); print(f"  CVM: {len(fiduc)} adm.fiduc | {len(gestor)} gestor | {len(custod)} custod")
    inst = universo_bcb(); print(f"  BCB: {len(inst)} instituições")
    porte = portes_bcb()
    gere = fundos_geridos(); cust_cnt = classes_custodiadas(); ini = inicio_atividade()
    print(f"  extra: {len(gere)} gestores c/ fundos | {len(cust_cnt)} custodiantes c/ classes | início p/ {len(ini)}")

    # marcas de grupos que JÁ têm DTVM ou administrador fiduciário (p/ pegar Vórtx SCD, etc.)
    infra_nomes = [x["nome"] for x in inst if "Distribuidora de TVM" in x["seg"]] + fiduc_nomes
    infra_brands = {b for b in (marca(n) for n in infra_nomes) if b and b not in GENERIC_BRAND}
    print(f"  marcas c/ DTVM/adm.fiduc.: {len(infra_brands)}")

    master = []
    for x in inst:
        raiz = so_dig(x["cnpj"]).zfill(8)[:8]
        af = raiz in fiduc
        p = porte.get(raiz, {})
        at = p.get("ativo"); esc = escala(p)
        cc = cust_cnt.get(raiz)
        mca = marca(x["nome"])
        grp = mca if (mca in infra_brands and not af and "Distribuidora de TVM" not in x["seg"]) else ""
        reg = {"Instituição": x["nome"], "Segmento BCB": x["seg"],
               "Perfil": perfil(x["nome"], x["seg"]), "Elegibilidade": elegibilidade(x["seg"], af),
               "Grupo tem DTVM/adm.fid.?": (f"sim — {grp}" if grp else ""),
               "CNPJ": fmt_cnpj(raiz), "UF": x["uf"], "Município": x["mun"],
               "Ativo Total (R$)": at or "", "Passivo Exigível (R$)": p.get("passivo") or "",
               "Patrimônio Líquido (R$)": p.get("pl") or "", "Captações (R$)": p.get("captacoes") or "",
               "Carteira de Crédito (R$)": p.get("carteira") or "", "TVM (R$)": p.get("tvm") or "",
               "Maior escala (R$)": esc or "", "Lucro Líquido (R$)": p.get("lucro") or "",
               "Início atividade": ini.get(raiz, ""),
               "Gestor CVM?": "sim" if raiz in gestor else "", "Nº fundos que gere": gere.get(raiz, "") or "",
               "Adm.fiduciário?": "sim" if af else "", "Custodiante?": "sim" if raiz in custod else "",
               "Nº classes que custodia": (cc[0] if cc else ""), "PL custodiado (R$)": (cc[1] if cc else ""),
               "Telefone": x["tel"], "E-mail": x["email"], "Site": x["site"],
               "_at": at or 0, "_esc": esc, "_el": elegibilidade(x["seg"], af), "_grp": grp}
        master.append(reg)

    # FILTRO DE PORTE (regra do usuário): nenhuma métrica de escala (ativo, passivo, PL, captações,
    # carteira, TVM) pode passar de R$ 400 mi. Escala confirmada acima do teto -> fora. Escala em branco
    # (não reportada) -> fica, EXCETO se o perfil já é sabidamente grande (Estrangeiro/Grande/Estatal/
    # Captivo), caso em que o branco é lacuna de dado, não pequenez -> fora (pega JP Morgan/BB/XP).
    def por_ativo(r): return (0, r["_at"]) if r["_at"] else (1, 0)   # com ativo primeiro, menor->maior
    def pequeno(r):
        if r["_esc"] > TETO: return False
        if r["_esc"] == 0 and r["Perfil"] in BIG_PERFIS: return False
        return True
    alvoA = sorted([r for r in master if r["_el"].startswith("Elegível direto") and not r["_grp"] and pequeno(r)], key=por_ativo)
    # Alvo B: financeiras (SCFI, elegíveis diretas à adm. desde a Res. 5.237/25) vêm PRIMEIRO;
    # depois as de objeto restrito (SCD/IP/câmbio/hipotecária), que só servem via DTVM nova.
    alvoB = sorted([r for r in master if (r["_el"].startswith("Elegível adm") or r["_el"].startswith("Só via DTVM"))
                    and not r["_grp"] and pequeno(r)],
                   key=lambda r: (0 if r["_el"].startswith("Elegível adm") else 1,) + por_ativo(r))
    master.sort(key=lambda r: (r["_el"], -(r["_at"] or 0)))

    xlsx = os.path.join(BASE, "bancos_alvo.xlsx")
    try: escrever(master, alvoA, alvoB, xlsx)
    except PermissionError:
        xlsx = os.path.join(BASE, "bancos_alvo_NOVO.xlsx"); escrever(master, alvoA, alvoB, xlsx)
        print("  (bancos_alvo.xlsx aberto — salvei como bancos_alvo_NOVO.xlsx)")
    with open(os.path.join(BASE, "bancos_alvo.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLS, extrasaction="ignore"); w.writeheader()
        for r in alvoA + alvoB: w.writerow(r)

    elA = [r for r in master if r["_el"].startswith("Elegível direto") and not r["_grp"]]
    elB = [r for r in master if (r["_el"].startswith("Elegível adm") or r["_el"].startswith("Só via DTVM")) and not r["_grp"]]
    cortadosA = sum(1 for r in elA if not pequeno(r)); cortadosB = sum(1 for r in elB if not pequeno(r))
    print("\n===== RESUMO =====")
    print("  por elegibilidade:", dict(collections.Counter(r["_el"] for r in master)))
    print("  por perfil:", dict(collections.Counter(r["Perfil"] for r in master)))
    print(f"  FILTRO DE PORTE (teto R$ {TETO/1e6:,.0f} mi por métrica de escala):")
    print(f"    Alvo A: {len(elA)} elegíveis -> {len(alvoA)} dentro do teto ({cortadosA} cortados por escala > teto)")
    print(f"    Alvo B: {len(elB)} elegíveis -> {len(alvoB)} dentro do teto ({cortadosB} cortados por escala > teto)")
    print("\n  Alvo A — top 15 menores por Ativo Total (Perfil):")
    for r in alvoA[:15]:
        at = f"{r['_at']/1e6:,.0f} mi" if r["_at"] else "?"
        print(f"    {r['Instituição'][:40]:40} | {r['Perfil'][:20]:20} | {r['UF']:2} | Ativo R$ {at}")

if __name__ == "__main__":
    main()
